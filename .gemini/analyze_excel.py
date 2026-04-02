import openpyxl

wb = openpyxl.load_workbook('Calculo_escala.xlsx', data_only=False)
sheet = wb.active

print(f"Sheet Name: {sheet.title}")
print("--- Cells with Formulas ---")
for row in sheet.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"{cell.coordinate}: {cell.value}")

print("\n--- Cells with Data (Inputs/Constants) ---")
for row in sheet.iter_rows():
    for cell in row:
        if cell.value and not (isinstance(cell.value, str) and cell.value.startswith('=')):
            print(f"{cell.coordinate}: {cell.value}")
