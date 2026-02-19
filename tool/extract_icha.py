import openpyxl
import json
import re

wb = openpyxl.load_workbook(
    r'c:\Users\andre\Desktop\WebProject\tool\ICHA_PROFILES.xlsx', data_only=True)


def safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(',', '.').replace(' ', '')
    if s in ['-', '', 'None']:
        return None
    try:
        return float(s)
    except:
        return None


def extract_sheet(ws, start_row, designation_col, weight_col, data_cols, designation_prefix=""):
    """Generic extraction for ICHA sheets.
    data_cols: list of (col_index, property_name) pairs
    """
    profiles = []
    current_desig_base = ""

    for r in range(start_row, ws.max_row + 1):
        # Check if row has weight data
        w = safe_float(ws.cell(r, weight_col).value)
        if w is None:
            continue

        # Check designation column for new group
        desig_val = ws.cell(r, designation_col).value
        if desig_val is not None:
            desig_str = str(desig_val).strip()
            if desig_str and desig_str not in ['cm *', 'cm*', 'cm * ', 'cm*cm', 'cm *cm', 'cm *cm*']:
                current_desig_base = desig_str

        if not current_desig_base:
            continue

        profile = {"weight": w}

        # Build designation name
        # Clean up designation base
        base = current_desig_base.replace('  ', ' ').strip()
        if base.endswith('*'):
            base = base[:-1].strip()

        # Read dimension columns for designation
        for col_idx, prop_name in data_cols:
            val = safe_float(ws.cell(r, col_idx).value)
            if val is not None:
                profile[prop_name] = val

        # Build the full designation string
        profile["designation"] = f"{base}x{w}"
        profile["base"] = base

        if len(profile) > 3:  # Must have at least some properties
            profiles.append(profile)

    return profiles

# ===== EXTRACT EACH SERIES =====


all_series = {}

# --- L Alas Iguales ---
ws = wb['Alas Iguales']
profiles = []
current_base = ""
for r in range(11, 53):
    w = safe_float(ws.cell(r, 3).value)
    if w is None:
        continue
    d = ws.cell(r, 2).value
    if d and str(d).strip() not in ['cm *', 'cm*']:
        current_base = str(d).strip().rstrip('*').strip()
    if not current_base:
        continue
    p = {"designation": f"{current_base}x{w}",
         "base": current_base, "weight": w}
    for col, name in [(4, 'B_mm'), (5, 'e_mm'), (6, 'A_cm2'), (7, 'Ix_cm4'), (8, 'Wx_cm3'), (9, 'ix_cm'), (10, 'xy_cm'), (11, 'iu_cm'), (12, 'iv_cm')]:
        v = safe_float(ws.cell(r, col).value)
        if v is not None:
            p[name] = v
    profiles.append(p)
all_series["L"] = {"name": "Ángulo Alas Iguales",
                   "icon": "fa-solid fa-l", "profiles": profiles}

# --- L Alas Desiguales ---
ws = wb['Alas Desiguales']
profiles = []
current_base = ""
for r in range(11, 48):
    w = safe_float(ws.cell(r, 3).value)
    if w is None:
        continue
    d = ws.cell(r, 2).value
    if d and str(d).strip() not in ['cm *cm', 'cm*cm', 'cm *cm*']:
        current_base = str(d).strip().rstrip('*').strip()
    if not current_base:
        continue
    p = {"designation": f"{current_base}x{w}",
         "base": current_base, "weight": w}
    for col, name in [(4, 'e_mm'), (5, 'A_cm2'), (6, 'Ix_cm4'), (7, 'Wx_cm3'), (8, 'ix_cm'), (9, 'y_cm'), (10, 'Iy_cm4'), (11, 'Wy_cm3'), (12, 'iy_cm'), (13, 'x_cm'), (14, 'iu_cm'), (15, 'iv_cm')]:
        v = safe_float(ws.cell(r, col).value)
        if v is not None:
            p[name] = v
    profiles.append(p)
all_series["L_desig"] = {"name": "Ángulo Alas Desiguales",
                         "icon": "fa-solid fa-l", "profiles": profiles}

# --- TL Alas Iguales ---
ws = wb['TL Alas iguales']
profiles = []
current_base = ""
for r in range(12, 54):
    w = safe_float(ws.cell(r, 3).value)
    if w is None:
        continue
    d = ws.cell(r, 2).value
    if d and str(d).strip() not in ['cm*', 'cm *']:
        current_base = str(d).strip().rstrip('*').strip()
    if not current_base:
        continue
    p = {"designation": f"{current_base}x{w}",
         "base": current_base, "weight": w}
    for col, name in [(4, 'B_mm'), (5, 'e_mm'), (6, 'A_cm2'), (7, 'ix_cm')]:
        v = safe_float(ws.cell(r, col).value)
        if v is not None:
            p[name] = v
    profiles.append(p)
all_series["TL"] = {
    "name": "Doble Ángulo (TL)", "icon": "fa-solid fa-t", "profiles": profiles}

# --- XL Alas Iguales ---
ws = wb['XL Alas iguales']
profiles = []
current_base = ""
for r in range(11, 53):
    w = safe_float(ws.cell(r, 3).value)
    if w is None:
        continue
    d = ws.cell(r, 2).value
    if d and str(d).strip() not in ['cm*', 'cm *', 'TL  H* peso ']:
        current_base = str(d).strip().rstrip('*').strip()
    if not current_base:
        continue
    p = {"designation": f"{current_base}x{w}",
         "base": current_base, "weight": w}
    for col, name in [(4, 'B_mm'), (5, 'e_mm'), (6, 'A_cm2'), (7, 'iv_cm')]:
        v = safe_float(ws.cell(r, col).value)
        if v is not None:
            p[name] = v
    profiles.append(p)
all_series["XL"] = {
    "name": "Ángulo en Cruz (XL)", "icon": "fa-solid fa-xmark", "profiles": profiles}

# --- Serie C (Diseño) ---
ws = wb['Serie C (Diseño)']
profiles = []
current_base = ""
for r in range(11, 72):
    w = safe_float(ws.cell(r, 3).value)
    if w is None:
        continue
    d = ws.cell(r, 2).value
    if d and str(d).strip() not in ['cm *', 'cm*']:
        current_base = str(d).strip().rstrip('*').strip()
    if not current_base:
        continue
    p = {"designation": f"{current_base}x{w}",
         "base": current_base, "weight": w}
    for col, name in [(4, 'B_mm'), (5, 'e_mm'), (6, 'A_cm2'), (7, 'Ix_cm4'), (8, 'Wx_cm3'), (9, 'ix_cm'), (10, 'Iy_cm4'), (11, 'Wy_cm3'), (12, 'iy_cm'), (13, 'X_cm')]:
        v = safe_float(ws.cell(r, col).value)
        if v is not None:
            p[name] = v
    profiles.append(p)
all_series["C"] = {"name": "Canal C",
                   "icon": "fa-solid fa-c", "profiles": profiles}

# --- Serie IC (Diseño) ---
ws = wb['Serie IC (Diseño)']
profiles = []
current_base = ""
for r in range(12, 86):
    w = safe_float(ws.cell(r, 3).value)
    if w is None:
        continue
    d = ws.cell(r, 2).value
    if d and str(d).strip() not in ['cm *', 'cm*']:
        current_base = str(d).strip().rstrip('*').strip()
    if not current_base:
        continue
    p = {"designation": f"{current_base}x{w}",
         "base": current_base, "weight": w}
    for col, name in [(4, 'B_mm'), (5, 'e_mm'), (6, 'A_cm2'), (7, 'Iy_cm4'), (8, 'Wy_cm3')]:
        v = safe_float(ws.cell(r, col).value)
        if v is not None:
            p[name] = v
    profiles.append(p)
all_series["IC"] = {
    "name": "Doble Canal (IC)", "icon": "fa-solid fa-pause", "profiles": profiles}

# --- Serie CA (Diseño) ---
ws = wb['Serie CA (Diseño)']
profiles = []
current_base = ""
for r in range(11, 52):
    w = safe_float(ws.cell(r, 3).value)
    if w is None:
        continue
    d = ws.cell(r, 2).value
    if d and str(d).strip() not in ['cm *', 'cm*']:
        current_base = str(d).strip().rstrip('*').strip()
    if not current_base:
        continue
    p = {"designation": f"{current_base}x{w}",
         "base": current_base, "weight": w}
    for col, name in [(4, 'B_mm'), (5, 'C_mm'), (6, 'e_mm'), (7, 'A_cm2'), (8, 'Ix_cm4'), (9, 'Wx_cm3'), (10, 'ix_cm'), (11, 'Iy_cm4'), (12, 'Wy_cm3'), (13, 'iy_cm'), (14, 'X_cm')]:
        v = safe_float(ws.cell(r, col).value)
        if v is not None:
            p[name] = v
    profiles.append(p)
all_series["CA"] = {
    "name": "Canal Atiesada (CA)", "icon": "fa-solid fa-c", "profiles": profiles}

# --- Serie ICA (Diseño) ---
ws = wb['Serie ICA (Diseño)']
profiles = []
current_base = ""
for r in range(12, 53):
    w = safe_float(ws.cell(r, 3).value)
    if w is None:
        continue
    d = ws.cell(r, 2).value
    if d and str(d).strip() not in ['cm *', 'cm*']:
        current_base = str(d).strip().rstrip('*').strip()
    if not current_base:
        continue
    p = {"designation": f"{current_base}x{w}",
         "base": current_base, "weight": w}
    for col, name in [(4, 'B_mm'), (5, 'C_mm'), (6, 'e_mm'), (7, 'A_cm2'), (8, 'Iy_cm4'), (9, 'Wy_cm3')]:
        v = safe_float(ws.cell(r, col).value)
        if v is not None:
            p[name] = v
    profiles.append(p)
all_series["ICA"] = {
    "name": "Doble Canal Atiesada (ICA)", "icon": "fa-solid fa-pause", "profiles": profiles}

# --- Serie Cajón ---
ws = wb['Serie Cajon']
profiles = []
current_base = ""
for r in range(11, 56):
    w = safe_float(ws.cell(r, 3).value)
    if w is None:
        continue
    d = ws.cell(r, 2).value
    if d and str(d).strip() not in ['cm *', 'cm*']:
        current_base = str(d).strip().rstrip('*').strip()
    if not current_base:
        continue
    p = {"designation": f"{current_base}x{w}",
         "base": current_base, "weight": w}
    for col, name in [(4, 'e_mm'), (5, 'A_cm2'), (6, 'Ix_cm4'), (7, 'Wx_cm3'), (8, 'ix_cm'), (9, 'Iy_cm4'), (10, 'Wy_cm3'), (11, 'iy_cm')]:
        v = safe_float(ws.cell(r, col).value)
        if v is not None:
            p[name] = v
    profiles.append(p)
all_series["CAJON"] = {"name": "Cajón Rectangular",
                       "icon": "fa-regular fa-square", "profiles": profiles}

# --- Serie IN (Diseño) ---
ws = wb['Serie IN (Diseño)']
profiles = []
current_base = ""
for r in range(11, 170):
    w = safe_float(ws.cell(r, 3).value)
    if w is None:
        continue
    d = ws.cell(r, 2).value
    if d and str(d).strip() not in ['cm *', 'cm*']:
        current_base = str(d).strip().rstrip('*').strip()
    if not current_base:
        continue
    p = {"designation": f"{current_base}x{w}",
         "base": current_base, "weight": w}
    for col, name in [(4, 'B_mm'), (5, 'e_mm'), (6, 't_mm'), (7, 'A_cm2'), (8, 'Ix_cm4'), (9, 'Wx_cm3'), (10, 'ix_cm'), (11, 'Iy_cm4'), (12, 'Wy_cm3'), (13, 'iy_cm')]:
        v = safe_float(ws.cell(r, col).value)
        if v is not None:
            p[name] = v
    profiles.append(p)
all_series["IN"] = {"name": "Viga Soldada IN",
                    "icon": "fa-solid fa-i", "profiles": profiles}

# --- Serie IP (Diseño) ---
if 'Serie IP (Diseño)' in wb.sheetnames:
    ws = wb['Serie IP (Diseño)']
    profiles = []
    current_base = ""
    for r in range(11, 200):
        w = safe_float(ws.cell(r, 3).value)
        if w is None:
            continue
        d = ws.cell(r, 2).value
        if d and str(d).strip() not in ['cm *', 'cm*']:
            current_base = str(d).strip().rstrip('*').strip()
        if not current_base:
            continue
        p = {"designation": f"{current_base}x{w}",
             "base": current_base, "weight": w}
        for col, name in [(4, 'B_mm'), (5, 'e_mm'), (6, 't_mm'), (7, 'A_cm2'), (8, 'Ix_cm4'), (9, 'Wx_cm3'), (10, 'ix_cm'), (11, 'Iy_cm4'), (12, 'Wy_cm3'), (13, 'iy_cm')]:
            v = safe_float(ws.cell(r, col).value)
            if v is not None:
                p[name] = v
        profiles.append(p)
    if profiles:
        all_series["IP"] = {"name": "Viga Soldada IP",
                            "icon": "fa-solid fa-i", "profiles": profiles}

# --- Serie HN (Diseño) ---
if 'Serie HN (Diseño)' in wb.sheetnames:
    ws = wb['Serie HN (Diseño)']
    profiles = []
    current_base = ""
    for r in range(11, 200):
        w = safe_float(ws.cell(r, 3).value)
        if w is None:
            continue
        d = ws.cell(r, 2).value
        if d and str(d).strip() not in ['cm *', 'cm*']:
            current_base = str(d).strip().rstrip('*').strip()
        if not current_base:
            continue
        p = {"designation": f"{current_base}x{w}",
             "base": current_base, "weight": w}
        for col, name in [(4, 'B_mm'), (5, 'e_mm'), (6, 't_mm'), (7, 'A_cm2'), (8, 'Ix_cm4'), (9, 'Wx_cm3'), (10, 'ix_cm'), (11, 'Iy_cm4'), (12, 'Wy_cm3'), (13, 'iy_cm')]:
            v = safe_float(ws.cell(r, col).value)
            if v is not None:
                p[name] = v
        profiles.append(p)
    if profiles:
        all_series["HN"] = {"name": "Perfil HN",
                            "icon": "fa-solid fa-h", "profiles": profiles}

# --- Serie PH (Diseño) ---
if 'Serie PH (Diseño)' in wb.sheetnames:
    ws = wb['Serie PH (Diseño)']
    profiles = []
    current_base = ""
    for r in range(11, 200):
        w = safe_float(ws.cell(r, 3).value)
        if w is None:
            continue
        d = ws.cell(r, 2).value
        if d and str(d).strip() not in ['cm *', 'cm*']:
            current_base = str(d).strip().rstrip('*').strip()
        if not current_base:
            continue
        p = {"designation": f"{current_base}x{w}",
             "base": current_base, "weight": w}
        for col, name in [(4, 'B_mm'), (5, 'e_mm'), (6, 't_mm'), (7, 'A_cm2'), (8, 'Ix_cm4'), (9, 'Wx_cm3'), (10, 'ix_cm'), (11, 'Iy_cm4'), (12, 'Wy_cm3'), (13, 'iy_cm')]:
            v = safe_float(ws.cell(r, col).value)
            if v is not None:
                p[name] = v
        profiles.append(p)
    if profiles:
        all_series["PH"] = {"name": "Perfil PH",
                            "icon": "fa-solid fa-h", "profiles": profiles}

# ===== GENERATE JS FILE =====
js_content = "// ICHA Profile Catalog Data\n"
js_content += "// Auto-generated from ICHA_PROFILES.xlsx\n"
js_content += f"// Total series: {len(all_series)}\n"
total_profiles = sum(len(s['profiles']) for s in all_series.values())
js_content += f"// Total profiles: {total_profiles}\n\n"
js_content += "const ICHA_CATALOG = " + \
    json.dumps(all_series, ensure_ascii=False, indent=2) + ";\n"

with open(r'c:\Users\andre\Desktop\WebProject\tool\icha_data.js', 'w', encoding='utf-8') as f:
    f.write(js_content)

print(f"Generated icha_data.js")
print(f"Series: {list(all_series.keys())}")
for k, v in all_series.items():
    print(f"  {k}: {v['name']} - {len(v['profiles'])} profiles")
print(f"Total profiles: {total_profiles}")
