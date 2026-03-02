import openpyxl
import os

excel_file = r'c:\Users\andre\Desktop\WebProject\Acortadores de Pandeo - Perfil XL Laminado.xlsx'
output_file = r'c:\Users\andre\Desktop\WebProject\tool\buckling_analysis.txt'

def analyze_excel():
    if not os.path.exists(excel_file):
        print(f"Error: {excel_file} not found.")
        return

    # Load with formulas
    wb_formulas = openpyxl.load_workbook(excel_file, data_only=False)
    # Load with values
    wb_values = openpyxl.load_workbook(excel_file, data_only=True)

    with open(output_file, 'w', encoding='utf-8') as out:
        out.write(f"Analysis of {os.path.basename(excel_file)}\n")
        out.write(f"Sheet names: {wb_formulas.sheetnames}\n\n")

        for sname in wb_formulas.sheetnames:
            ws_f = wb_formulas[sname]
            ws_v = wb_values[sname]
            
            out.write(f"\n{'='*80}\n")
            out.write(f"SHEET: {sname}\n")
            out.write(f"{'='*80}\n")

            for r in range(1, ws_f.max_row + 1):
                row_data = []
                has_content = False
                for c in range(1, ws_f.max_column + 1):
                    val = ws_v.cell(r, c).value
                    formula = ws_f.cell(r, c).value
                    
                    if val is not None or formula is not None:
                        has_content = True
                    
                    if formula != val and isinstance(formula, str) and formula.startswith('='):
                        row_data.append(f"[{val} (F: {formula})]")
                    else:
                        row_data.append(str(val) if val is not None else "")

                if has_content:
                    # Trim trailing empty cells
                    while row_data and row_data[-1] == "":
                        row_data.pop()
                    if row_data:
                        out.write(f"R{r}: {' | '.join(row_data)}\n")

    print(f"Analysis complete. Results written to {output_file}")

if __name__ == "__main__":
    analyze_excel()
