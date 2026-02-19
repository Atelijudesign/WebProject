import openpyxl

wb = openpyxl.load_workbook(
    r'c:\Users\andre\Desktop\WebProject\tool\ICHA_PROFILES.xlsx', data_only=True)

with open(r'c:\Users\andre\Desktop\WebProject\tool\excel_dump.txt', 'w', encoding='utf-8') as f:
    f.write(f"Sheets: {wb.sheetnames}\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f"\n{'='*80}\n")
        f.write(f"Sheet: {sheet_name}\n")
        f.write(f"Dimensions: {ws.dimensions}\n")
        f.write(f"Max row: {ws.max_row}, Max col: {ws.max_column}\n")

        # Print merged cells
        merged = list(ws.merged_cells.ranges)
        f.write(f"Merged cells count: {len(merged)}\n")
        for mc in merged[:20]:
            f.write(f"  Merged: {mc}\n")

        # Print ALL rows with data
        f.write(f"\n--- ALL NON-EMPTY ROWS ---\n")
        for r in range(1, ws.max_row + 1):
            row_data = []
            for c in range(1, min(ws.max_column + 1, 26)):
                v = ws.cell(r, c).value
                if v is not None:
                    row_data.append(f"C{c}={v}")
            if row_data:
                f.write(f"  R{r}: {' | '.join(row_data)}\n")

print("Done! Output written to excel_dump.txt")
