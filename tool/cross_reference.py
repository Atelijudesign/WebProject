"""
Cross-reference ICHA profile data between Excel and icha_data.js.
Finds discrepancies in mechanical properties.
"""
import openpyxl
import json
import re
import sys

# ---- 1. Read Excel data ----
wb = openpyxl.load_workbook(
    r'c:\Users\andre\Desktop\WebProject\tool\ICHA_PROFILES.xlsx')


def read_sheet_profiles(ws, config):
    """Read profiles from a sheet given a configuration dict."""
    profiles = []
    header_row = config.get('header_row', 10)  # row with units
    data_start = config.get('data_start', 11)
    # dict mapping property_name -> column_index (1-based)
    cols = config['cols']

    current_designation = None
    for r in range(data_start, ws.max_row + 1):
        # Check if row has data
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(v is not None for v in row_vals):
            continue

        # Get designation cell
        desig_col = cols.get('designation', 2)
        desig_val = ws.cell(r, desig_col).value
        if desig_val is not None and str(desig_val).strip():
            current_designation = str(desig_val).strip()

        if current_designation is None:
            continue

        # Get weight
        weight_col = cols.get('weight', 3)
        weight = ws.cell(r, weight_col).value
        if weight is None or not isinstance(weight, (int, float)):
            continue

        profile = {'designation': current_designation, 'weight': float(weight)}

        # Read all configured columns
        for prop, col in cols.items():
            if prop in ('designation', 'weight'):
                continue
            val = ws.cell(r, col).value
            if val is not None and isinstance(val, (int, float)):
                profile[prop] = float(val)
            elif val is not None and isinstance(val, str):
                # Try to parse numeric strings like "5 ,29" or "3,7 8"
                cleaned = val.replace(' ', '').replace(',', '.')
                try:
                    profile[prop] = float(cleaned)
                except ValueError:
                    pass  # Skip non-numeric

        profiles.append(profile)

    return profiles


# Sheet configurations - mapping property names to column indices
# Column indices are 1-based

# L Alas Iguales: Designation=B2, Weight=C, B_mm=D, e_mm=E, A_cm2=F, Ix_cm4=G, Wx_cm3=H, ix_cm=I, xy_cm=J, iu_cm=K, iv_cm=L
SHEET_CONFIGS = {
    'Alas Iguales': {
        'series': 'L',
        'cols': {
            'designation': 2, 'weight': 3,
            'B_mm': 4, 'e_mm': 5, 'A_cm2': 6,
            'Ix_cm4': 7, 'Wx_cm3': 8, 'ix_cm': 9, 'xy_cm': 10,
            'iu_cm': 11, 'iv_cm': 12
        }
    },
    'Alas Desiguales': {
        'series': 'L_desig',
        'cols': {
            'designation': 2, 'weight': 3,
            'e_mm': 4, 'A_cm2': 5,
            'Ix_cm4': 6, 'Wx_cm3': 7, 'ix_cm': 8, 'y_cm': 9,
            'Iy_cm4': 10, 'Wy_cm3': 11, 'iy_cm': 12, 'x_cm': 13,
            'iu_cm': 14, 'iv_cm': 15
        }
    },
    'Serie C (Diseño)': {
        'series': 'C',
        'cols': {
            'designation': 2, 'weight': 3,
            'B_mm': 4, 'e_mm': 5, 'A_cm2': 6,
            'Ix_cm4': 7, 'Wx_cm3': 8, 'ix_cm': 9,
            'Iy_cm4': 10, 'Wy_cm3': 11, 'iy_cm': 12, 'X_cm': 13
        }
    },
    'Serie CA (Diseño)': {
        'series': 'CA',
        'cols': {
            'designation': 2, 'weight': 3,
            'B_mm': 4, 'C_mm': 5, 'e_mm': 6, 'A_cm2': 7,
            'Ix_cm4': 8, 'Wx_cm3': 9, 'ix_cm': 10,
            'Iy_cm4': 11, 'Wy_cm3': 12, 'iy_cm': 13, 'X_cm': 14
        }
    },
    'Serie IN (Diseño)': {
        'series': 'IN',
        'cols': {
            'designation': 2, 'weight': 3,
            'B_mm': 4, 'e_mm': 5, 't_mm': 6, 'A_cm2': 7,
            'Ix_cm4': 8, 'Wx_cm3': 9, 'ix_cm': 10,
            'Iy_cm4': 11, 'Wy_cm3': 12, 'iy_cm': 13
        }
    },
    'Serie IP (Diseño)': {
        'series': 'IP',
        'cols': {
            'designation': 2, 'weight': 3,
            'B_mm': 4, 'e_mm': 5, 't_mm': 6, 'A_cm2': 7,
            'Ix_cm4': 8, 'Wx_cm3': 9, 'ix_cm': 10,
            'Iy_cm4': 11, 'Wy_cm3': 12, 'iy_cm': 13
        }
    },
    'Serie HN (Diseño)': {
        'series': 'HN',
        'cols': {
            'designation': 2, 'weight': 3,
            'B_mm': 4, 'e_mm': 5, 't_mm': 6, 'A_cm2': 7,
            'Ix_cm4': 8, 'Wx_cm3': 9, 'ix_cm': 10,
            'Iy_cm4': 11, 'Wy_cm3': 12, 'iy_cm': 13
        }
    },
    'Serie PH (Diseño)': {
        'series': 'PH',
        'cols': {
            'designation': 2, 'weight': 3,
            'B_mm': 4, 'e_mm': 5, 't_mm': 6, 'A_cm2': 7,
            'Ix_cm4': 8, 'Wx_cm3': 9, 'ix_cm': 10,
            'Iy_cm4': 11, 'Wy_cm3': 12, 'iy_cm': 13
        }
    },
}

# Read Excel profiles
excel_data = {}
for sheet_name, config in SHEET_CONFIGS.items():
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        profiles = read_sheet_profiles(ws, config)
        series = config['series']
        excel_data[series] = profiles
        print(f"Excel: {sheet_name} -> {series}: {len(profiles)} profiles")

# ---- 2. Read JS data ----
js_path = r'c:\Users\andre\Desktop\WebProject\tool\icha_data.js'
with open(js_path, 'r', encoding='utf-8') as f:
    content = f.read()

# Extract the JSON-like object from the JS file
# Find the assignment: const ICHA_CATALOG = { ... }
match = re.search(
    r'const\s+ICHA_CATALOG\s*=\s*(\{.*\})\s*;?\s*$', content, re.DOTALL)
if not match:
    print("ERROR: Could not find ICHA_CATALOG in the JS file")
    sys.exit(1)

json_str = match.group(1)
catalog = json.loads(json_str)

js_data = {}
for series_key, series_obj in catalog.items():
    js_data[series_key] = series_obj.get('profiles', [])
    print(
        f"JS: {series_key} ({series_obj.get('name', '?')}): {len(js_data[series_key])} profiles")

# ---- 3. Cross-reference ----
out = open(r'c:\Users\andre\Desktop\WebProject\tool\cross_reference_report.txt',
           'w', encoding='utf-8')

# Map Excel series to JS series
SERIES_MAP = {
    'L': 'L',
    'L_desig': 'L_desig',
    'C': 'C',
    'CA': 'CA',
    'IN': 'IN',
    'IP': 'IP',
    'HN': 'HN',
    'PH': 'PH',
}

total_compared = 0
total_discrepancies = 0
total_missing = 0

for excel_series, js_series in SERIES_MAP.items():
    if excel_series not in excel_data:
        out.write(f"\n{'='*60}\n")
        out.write(f"SKIPPED: Excel series '{excel_series}' not loaded\n")
        continue

    if js_series not in js_data:
        out.write(f"\n{'='*60}\n")
        out.write(f"SKIPPED: JS series '{js_series}' not found\n")
        continue

    out.write(f"\n{'='*60}\n")
    out.write(f"COMPARING: Excel '{excel_series}' vs JS '{js_series}'\n")
    out.write(f"{'='*60}\n")

    excel_profiles = excel_data[excel_series]
    js_profiles = js_data[js_series]

    # Build JS lookup by weight (since weight is the unique identifier within a base)
    # We match by weight since designations may differ
    for ep in excel_profiles:
        weight = ep['weight']

        # Find matching JS profile by weight
        matches = [jp for jp in js_profiles if abs(
            jp.get('weight', 0) - weight) < 0.05]

        if not matches:
            out.write(
                f"\n  MISSING in JS: {ep.get('designation', '?')} w={weight}\n")
            total_missing += 1
            continue

        jp = matches[0]
        total_compared += 1

        # Compare all common properties
        props_to_compare = ['A_cm2', 'B_mm', 'e_mm', 't_mm', 'C_mm',
                            'Ix_cm4', 'Wx_cm3', 'ix_cm',
                            'Iy_cm4', 'Wy_cm3', 'iy_cm',
                            'iu_cm', 'iv_cm', 'xy_cm',
                            'x_cm', 'y_cm', 'X_cm']

        discreps = []
        for prop in props_to_compare:
            if prop in ep and prop in jp:
                ev = ep[prop]
                jv = jp[prop]
                # Allow tolerance
                if abs(ev - jv) > 0.5 and abs(ev - jv) / max(abs(ev), 1) > 0.005:
                    discreps.append(
                        f"{prop}: Excel={ev} JS={jv} diff={abs(ev-jv):.2f}")

        if discreps:
            total_discrepancies += 1
            out.write(
                f"\n  DISCREPANCY: {ep.get('designation', '?')} w={weight} (JS: {jp.get('designation', '?')})\n")
            for d in discreps:
                out.write(f"    {d}\n")

out.write(f"\n\n{'='*60}\n")
out.write(f"SUMMARY\n")
out.write(f"{'='*60}\n")
out.write(f"Total profiles compared: {total_compared}\n")
out.write(f"Total with discrepancies: {total_discrepancies}\n")
out.write(f"Total missing from JS: {total_missing}\n")

out.close()
print(f"\nDone! Report written to cross_reference_report.txt")
print(
    f"Compared: {total_compared}, Discrepancies: {total_discrepancies}, Missing: {total_missing}")
