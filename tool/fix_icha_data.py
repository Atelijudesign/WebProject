"""
Script to fix icha_data.js by reading ICHA_PROFILES.xlsx with proper value inheritance
and patching the JSON data.
"""
import openpyxl
import json
import re
import sys

# ---- 1. Read Excel data WITH INHERITANCE ----
wb = openpyxl.load_workbook(
    r'c:\Users\andre\Desktop\WebProject\tool\ICHA_PROFILES.xlsx')

SHEET_CONFIGS = {
    'Alas Iguales': {
        'series': 'L',
        'cols': {
            'designation': 2, 'weight': 3,
            'B_mm': 4, 'e_mm': 5, 'A_cm2': 6,
            'Ix_cm4': 7, 'Wx_cm3': 8, 'ix_cm': 9, 'xy_cm': 10,
            'iu_cm': 11, 'iv_cm': 12
        }
    },
    'Alas Desiguales': {
        'series': 'L_desig',
        'cols': {
            'designation': 2, 'weight': 3,
            'e_mm': 4, 'A_cm2': 5,
            'Ix_cm4': 6, 'Wx_cm3': 7, 'ix_cm': 8, 'y_cm': 9,
            'Iy_cm4': 10, 'Wy_cm3': 11, 'iy_cm': 12, 'x_cm': 13,
            'iu_cm': 14, 'iv_cm': 15
        }
    },
    'Serie C (Diseño)': {
        'series': 'C',
        'cols': {
            'designation': 2, 'weight': 3,
            'B_mm': 4, 'e_mm': 5, 'A_cm2': 6,
            'Ix_cm4': 7, 'Wx_cm3': 8, 'ix_cm': 9,
            'Iy_cm4': 10, 'Wy_cm3': 11, 'iy_cm': 12, 'X_cm': 13
        }
    },
    'Serie CA (Diseño)': {
        'series': 'CA',
        'cols': {
            'designation': 2, 'weight': 3,
            'B_mm': 4, 'C_mm': 5, 'e_mm': 6, 'A_cm2': 7,
            'Ix_cm4': 8, 'Wx_cm3': 9, 'ix_cm': 10,
            'Iy_cm4': 11, 'Wy_cm3': 12, 'iy_cm': 13, 'X_cm': 14
        }
    },
    'Serie IN (Diseño)': {
        'series': 'IN',
        'cols': {
            'designation': 2, 'weight': 3,
            'B_mm': 4, 'e_mm': 5, 't_mm': 6, 'A_cm2': 7,
            'Ix_cm4': 8, 'Wx_cm3': 9, 'ix_cm': 10,
            'Iy_cm4': 11, 'Wy_cm3': 12, 'iy_cm': 13
        }
    },
    'Serie IP (Diseño)': {
        'series': 'IP',
        'cols': {
            'designation': 2, 'weight': 3,
            'B_mm': 4, 'e_mm': 5, 't_mm': 6, 'A_cm2': 7,
            'Ix_cm4': 8, 'Wx_cm3': 9, 'ix_cm': 10,
            'Iy_cm4': 11, 'Wy_cm3': 12, 'iy_cm': 13
        }
    },
    'Serie HN (Diseño)': {
        'series': 'HN',
        'cols': {
            'designation': 2, 'weight': 3,
            'B_mm': 4, 'e_mm': 5, 't_mm': 6, 'A_cm2': 7,
            'Ix_cm4': 8, 'Wx_cm3': 9, 'ix_cm': 10,
            'Iy_cm4': 11, 'Wy_cm3': 12, 'iy_cm': 13
        }
    },
    'Serie PH (Diseño)': {
        'series': 'PH',
        'cols': {
            'designation': 2, 'weight': 3,
            'B_mm': 4, 'e_mm': 5, 't_mm': 6, 'A_cm2': 7,
            'Ix_cm4': 8, 'Wx_cm3': 9, 'ix_cm': 10,
            'Iy_cm4': 11, 'Wy_cm3': 12, 'iy_cm': 13
        }
    },
}

excel_data = {}

for sheet_name, config in SHEET_CONFIGS.items():
    if sheet_name not in wb.sheetnames:
        continue

    ws = wb[sheet_name]
    cols = config['cols']
    series = config['series']

    print(f"Processing {sheet_name} ({series})...")

    rows_data = []
    # Values to carry forward
    last_values = {k: None for k in cols.keys()}
    current_base = None

    for r in range(11, ws.max_row + 1):
        # Check if row has data (weight column must be present)
        weight_col = cols.get('weight', 3)
        weight_val = ws.cell(r, weight_col).value

        if weight_val is None or not isinstance(weight_val, (int, float)):
            continue

        weight = float(weight_val)

        # Check designation (base)
        desig_col = cols.get('designation', 2)
        desig_val = ws.cell(r, desig_col).value

        if desig_val is not None and str(desig_val).strip():
            current_base = str(desig_val).strip()
            # Reset last values on new base?
            # Actually, sometimes B implies inheritance from top of block.
            # But usually a new designation like "IN 100" starts a block.
            # Let's keep logic: if cell is empty, use last_value. If cell has value, update last_value.

        if current_base is None:
            continue

        profile = {'base': current_base, 'weight': weight}

        # Process properties with inheritance
        for prop, col_idx in cols.items():
            if prop in ('designation', 'weight'):
                continue

            val = ws.cell(r, col_idx).value

            # If value exists, update last_value and use it
            if val is not None:
                # Clean strings if necessary
                if isinstance(val, str):
                    try:
                        clean_val = float(val.replace(
                            ' ', '').replace(',', '.'))
                        last_values[prop] = clean_val
                    except ValueError:
                        pass  # keep old value if parse fails? or set None?
                elif isinstance(val, (int, float)):
                    last_values[prop] = float(val)

            # Use last_value if available
            if last_values[prop] is not None:
                profile[prop] = last_values[prop]

        rows_data.append(profile)

    excel_data[series] = rows_data
    print(f"  Loaded {len(rows_data)} profiles.")

# ---- 2. Patch icha_data.js ----
js_path = r'c:\Users\andre\Desktop\WebProject\tool\icha_data.js'
with open(js_path, 'r', encoding='utf-8') as f:
    js_content = f.read()

# Extract catalog
match = re.search(
    r'const\s+ICHA_CATALOG\s*=\s*(\{.*\})\s*;?\s*$', js_content, re.DOTALL)
if not match:
    print("Error: content not found")
    sys.exit(1)

catalog = json.loads(match.group(1))

patched_count = 0
for series_key, series_obj in catalog.items():
    if series_key not in excel_data:
        continue

    js_profiles = series_obj.get('profiles', [])
    excel_profiles = excel_data[series_key]

    for jp in js_profiles:
        # Match by weight (closest)
        weight = jp.get('weight', 0)

        # Find match
        match_ep = None
        for ep in excel_profiles:
            if abs(ep['weight'] - weight) < 0.05:
                match_ep = ep
                break

        if match_ep:
            # Update properties
            for k, v in match_ep.items():
                if k not in ('base', 'weight'):
                    # Update if missing or significantly different
                    if k not in jp or (isinstance(v, float) and abs(jp.get(k, 0) - v) > 0.1):
                        jp[k] = v
                        patched_count += 1

print(f"Patched {patched_count} properties.")

# Write back
new_json = json.dumps(catalog, indent=2, ensure_ascii=False)
new_content = f"// ICHA Profile Catalog Data\n// Auto-generated from ICHA_PROFILES.xlsx\n// Total series: {len(catalog)}\n// Total profiles: {sum(len(s['profiles']) for s in catalog.values())}\n\nconst ICHA_CATALOG = {new_json}"

with open(js_path, 'w', encoding='utf-8') as f:
    f.write(new_content)

print("Successfully updated icha_data.js")
